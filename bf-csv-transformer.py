##############PYTHON 3.13.2
brewfather_recipe = 'Schwarzbier.json'
excel_layout = 'brewsheet_empty.xlsx'

import json
import csv
from openpyxl import load_workbook


#change this file / filepath to the Brewfather json output
with open(brewfather_recipe, 'r') as file:
    data = json.load(file)

#############################helper functions for conversions
def conv_plato(sg):
    return(round((-1 * 616.868) + (1111.14 * sg) - (630.272 * sg**2) + (135.997 * sg**3), 1))

def conv_ebc(srm):
    return round(float(srm*1.97), 1)

#############################variable collection

##line 1 title
beerName = data["name"]

#line 2 facts
style = data['style']['name']
og = conv_plato(data["og"])
abv = round(data['abv'], 1)
fg = conv_plato(data["fg"])
color = conv_ebc(data["color"])
ibu = data["ibu"]

#line 5-11 ingredients list
malt_count = 0
for fermentable in data['data']['mashFermentables']:
    malt_count += 1
    exec("malt" + str(malt_count) + " = fermentable['name']")


hoplist = []
alphalist = []
for hop in data['hops']:
    if hop['name'] not in hoplist:
        hoplist.append(hop['name'])
        alphalist.append(hop['alpha'])
hop_count = 0
for hop in hoplist:
    hop_count += 1
    exec("hop" + str(hop_count) + " = hop")
alpha_count = 0
for hopAlpha in alphalist:
    alpha_count += 1
    exec("hopAlpha" + str(alpha_count) + " = hopAlpha")


yeast_count = 0
for yeast in data['yeasts']:
    yeast_count += 1
    exec("yeast" + str(yeast_count) + " = yeast['productId']")

#line 13-20 water profile


mashWaterAmount = data["water"]["mashWaterAmount"]
spargeWaterAmount = data["water"]["spargeWaterAmount"]
corrWaterAmount = 10

#mash
mashDilution = round(float(data['water']['dilutionAmount']/data['water']['totalAdjustments']['volume']),2)*100
mashCaCl2 = round(float(data["water"]['mashAdjustments']["calciumChloride"]),1)
mashCaSO4 = round(float(data["water"]['mashAdjustments']["calciumSulfate"]),1)
mashMgSO4 = round(float(data["water"]['mashAdjustments']["magnesiumSulfate"]),1)
mashNaCl = round(float(data["water"]['mashAdjustments']["sodiumChloride"]),1)
mashNaHCO3 = round(float(data["water"]['mashAdjustments']['sodiumBicarbonate']),1)
mashLA = round(float(data["water"]['mashAdjustments']['acids'][0]['amount']),1)

#sparge
spargeDilution = mashDilution
spargeCaCl2 = round(float(data["water"]['spargeAdjustments']["calciumChloride"]),2)
spargeCaSO4 = round(float(data["water"]['spargeAdjustments']["calciumSulfate"]),1)
spargeMgSO4 = round(float(data["water"]['spargeAdjustments']["magnesiumSulfate"]),1)
spargeNaCl = round(float(data["water"]['spargeAdjustments']["sodiumChloride"]),1)
spargeNaHCO3 = round(float(data["water"]['spargeAdjustments']['sodiumBicarbonate']),1)
spargeLA = round(float(data["water"]['spargeAdjustments']['acids'][0]['amount']),1)

#corr
corrDilution = mashDilution
corrCaCl2 = round((spargeCaCl2/spargeWaterAmount)*corrWaterAmount,1)
corrCaSO4 = round((spargeCaSO4/spargeWaterAmount)*corrWaterAmount,1)
corrMgSO4 = round((spargeMgSO4/spargeWaterAmount)*corrWaterAmount,1)
corrNaCl = round((spargeNaCl/spargeWaterAmount)*corrWaterAmount,1)
corrNaHCO3 = round((spargeNaHCO3/spargeWaterAmount)*corrWaterAmount,1)
corrLA = round((spargeLA/spargeWaterAmount)*corrWaterAmount,1)


#line 22-27 mash
mashpH = round(float(data['water']['mashPh']),2)
mashMaltAmount = round(float(data['data']["mashFermentablesAmount"]),1)
Gussführung = f"1:{round(mashWaterAmount/mashMaltAmount, 1)}"

mashStep_count = 0
for mashStep in data['mash']['steps']:
    mashStep_count += 1
    exec("mashStepDur" + str(mashStep_count) + " = mashStep['stepTime']")
    exec("mashStepTemp" + str(mashStep_count) + " = mashStep['stepTemp']")


#line 29-35 lauter 
"""none"""

#line 37-46 boil
preBoilGravity = conv_plato(data["preBoilGravity"])
preBoilVolume = round(data['equipment']['boilSize'],1)

postBoilGravity = og
postBoilVolume = round(data['equipment']["postBoilKettleVol"],1)

hopdose_count = 0
for hopdose in data['hops']:
    hopdose_count += 1
    exec("hopdose" + str(hopdose_count) + " = hopdose['name']")
    exec("hopdoseTime" + str(hopdose_count) + " = hopdose['time']")
    exec("hopdoseAmount" + str(hopdose_count) + " = hopdose['amount']")



#line 48-52 cool
whirlpoolTime = data["equipment"]["whirlpoolTime"]
wortTemp = data['fermentation']['steps'][0]['stepTemp']


#fermsheet
fermYeastAmount = data["yeasts"][0]["amount"]

fermStep_count = 0
for fermStep in data['fermentation']["steps"]:
    fermStep_count += 1
    pressureBar = 0
    if fermStep['pressure'] is not None:
        pressureBar = round(float(fermStep['pressure'])*0.0689476, 1)
    exec("fermStepName" + str(fermStep_count) + " = fermStep['name']")
    exec("fermStepTemp" + str(fermStep_count) + " = fermStep['stepTemp']")
    exec("fermStepPressure" + str(fermStep_count) + " = pressureBar")


#############################csv generation
#############################brewsheet
df = [
    [None,None,None,None,None,None,None,None,None,None,None,None,None], #ok
    [beerName ,None,None,None,"[BatchNo]",None,None,None,"[Date]",None,None,None,"Brewer:" ,None,None], #ok
    [style,None,og,"°P",abv,"%",fg,"°P",color,"EBC",ibu,"IBU",None,None,None], #ok
    ["ZUTATEN",None,"Prod","No",None,None,None,"alpha","Prod","No",None,None,None,"Prod","No"], #ok
    [globals().get('malt1', None), None, None, None,None, globals().get('hop1', None), None, globals().get('hopAlpha1', None), None, None, None, globals().get('yeast1', None), None],
    [globals().get('malt2', None), None, None, None,None, globals().get('hop2', None), None, globals().get('hopAlpha2', None), None, None, None, globals().get('yeast2', None), None],
    [globals().get('malt3', None), None, None, None,None, globals().get('hop3', None), None, globals().get('hopAlpha3', None), None, None, None, globals().get('yeast3', None), None],
    [globals().get('malt4', None), None, None, None,None, globals().get('hop4', None), None, globals().get('hopAlpha4', None), None, None, None, globals().get('yeast4', None), None],
    [globals().get('malt5', None), None, None, None,None, globals().get('hop5', None), None, globals().get('hopAlpha5', None), None, None, None, globals().get('yeast5', None), None],
    [globals().get('malt6', None), None, None, None,None, globals().get('hop6', None), None, globals().get('hopAlpha6', None), None, None, None, globals().get('yeast6', None), None],
    [globals().get('malt7', None), None, None, None,None, globals().get('hop7', None), None, globals().get('hopAlpha7', None), None, None, None, globals().get('yeast7', None), None],
    ["WASSER",None,None,None,None,None,None,None,None,None,None,None,None,"Resp: ",None], #ok
    ["Maische [l]",None,mashWaterAmount,None,None,"Nachguss [l]",None,spargeWaterAmount,None,None,None,"Korrektur [l]",None,corrWaterAmount,None],
    ["Verdünnung [%]",None,mashDilution,None,None,"Verdünnung [%]",None,spargeDilution,None,None,None,"Verdünnung [%]",None,corrDilution,None],
    ["CaCl2 (33%) [g]",None,mashCaCl2,None,None,"CaCl2 (33%) [g]",None,spargeCaCl2,None,None,None,"CaCl2 (33%) [g]",None,corrCaCl2,None],
    ["CaSO4 [g]",None,mashCaSO4,None,None,"CaSO4 [g]",None,spargeCaSO4,None,None,None,"CaSO4 [g]",None,corrCaSO4,None],
    ["MgSO4 [g]",None,mashMgSO4,None,None,"MgSO4 [g]",None,spargeMgSO4,None,None,None,"MgSO4 [g]",None,corrMgSO4,None],
    ["NaCl [g]",None,mashNaCl,None,None,"NaCl [g]",None,spargeNaCl,None,None,None,"NaCl [g]",None,corrNaCl,None],
    ["NaHCO3 [g]",None,mashNaHCO3,None,None,"NaHCO3 [g]",None,spargeNaHCO3,None,None,None,"NaHCO3 [g]",None,corrNaHCO3,None],
    ["Lactic Acid 80% [ml]",None,mashLA,None,None,"Lactic Acid 80% [ml]",None,spargeLA,None,None,None,"Lactic Acid 80% [ml]",None,corrLA,None, None],
    ["MAISCHE",None,"act","min","soll","max",None,"°C","min","Notes",None,None,None,"Resp:" ,None],
    ["pH (10mins) [pH]",None,None,mashpH - 0.1,mashpH,mashpH+0.1,"Step1",globals().get('mashStepTemp1', None),globals().get('mashStepDur1', None),None,None,None,None,None,None],
    ["Korrektur LA [ml]",None,None,0,0,50,"Step2",globals().get('mashStepTemp2', None),globals().get('mashStepDur2', None),None,None,None,None,None,None],
    [None,None,None,None,None,None,"Step3",globals().get('mashStepTemp3', None),globals().get('mashStepDur3', None),None,None,None,None,None,None],
    ["Wasser [l]",None,None,mashWaterAmount - 10,mashWaterAmount,160,"Step4",globals().get('mashStepTemp4', None),globals().get('mashStepDur4', None),None,None,None,None,None,None],
    ["Malz [kg]",None,None,None,mashMaltAmount,None,"Step5",globals().get('mashStepTemp5', None),globals().get('mashStepDur5', None),None,None,None,None,None,None],
    ["Gussführung",None,None,None,Gussführung,None,"Step6",globals().get('mashStepTemp6', None),globals().get('mashStepDur6', None),None,None,None,None,None,None],
    ["LÄUTERN",None,"act","min","soll","max",None,None,None,"Notes",None,None,None,"Resp:" ,None],
    ["Vorderwürze pH",None,None,mashpH - 0.1,mashpH,mashpH + 0.1,None,None,None,None,None,None,None,None,None],
    ["Vorderwürze °P",None,None,(og * 2)-3,og * 2,(og * 2)+3,None,None,None,None,None,None,None,None],
    ["Milchsäure",None,None,0,0,50,None,None,None,None,None,None,None,None,None],
    ["Nachgusswasser",None,None,spargeWaterAmount-20,spargeWaterAmount,spargeWaterAmount+30,None,None,None,None,None,None,None,None,None],
    ["Glattwasser pH",None,None,mashpH,mashpH +0.3,mashpH +0.5,None,None,None,None,None,None,None,None,None],
    ["Glattwasser °P",None,None,3.0,5.0,7.0,None,None,None,None,None,None,None,None,None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None], #ok
    ["KOCHEN",None,"act","min","soll","max","Sorte","Zeit","Menge","Notes",None,None,None,"Resp:" ,None],
    ["preBoil pH",None,None,mashpH-0.2,mashpH,mashpH+0.2,globals().get('hopdose1',None),globals().get('hopdoseTime1',None),globals().get('hopdoseAmount1',None),None,None,None,None,None,None],
    ["preBoil °P",None,None,preBoilGravity-0.1,preBoilGravity,preBoilGravity+0.1,globals().get('hopdose2',None),globals().get('hopdoseTime2',None),globals().get('hopdoseAmount2',None),None,None,None,None,None,None],
    ["preBoil Volume",None,None,preBoilVolume-20,preBoilVolume,preBoilVolume+20,globals().get('hopdose3',None),globals().get('hopdoseTime3',None),globals().get('hopdoseAmount3',None),None,None,None,None,None,None],
    ["preBoil LA [ml]",None,None,0,0,50,globals().get('hopdose4',None),globals().get('hopdoseTime4',None),globals().get('hopdoseAmount4',None),None,None,None,None,None,None],
    [None,None,None,None,None,None,globals().get('hopdose5',None),globals().get('hopdoseTime5',None),globals().get('hopdoseAmount5',None),None,None,None,None,None,None],
    ["postBoil pH",None,None,mashpH-0.5,mashpH-0.4,mashpH-0.3,globals().get('hopdose6',None),globals().get('hopdoseTime6',None),globals().get('hopdoseAmount6',None),None,None,None,None,None,None],
    ["postBoil °P",None,None,og-0.1,og,og+0.1,globals().get('hopdose7',None),globals().get('hopdoseTime7',None),globals().get('hopdoseAmount7',None),None,None,None,None,None,None],
    ["postBoil Volume",None,None,postBoilVolume - 10,postBoilVolume,postBoilVolume + 10,globals().get('hopdose8',None),globals().get('hopdoseTime8',None),globals().get('hopdoseAmount8',None),None,None,None,None,None,None],
    ["postBoil LA [ml]",None,None,0,0,50,globals().get('hopdose9',None),globals().get('hopdoseTime9',None),globals().get('hopdoseAmount9',None),None,None,None,None,None,None],
    [None,None,None,None,None,None,globals().get('hopdose10',None),globals().get('hopdoseTime10',None),globals().get('hopdoseAmount10',None),None,None,None,None,None,None],
    ["KÜHLEN",None,"act","min","soll","max",None,None,None,"Notes",None,None,None,"Resp:" ,None],
    ["Whirlpool [mins]",None,None,whirlpoolTime-5,whirlpoolTime,whirlpoolTime+5,None,None,None,None,None,None,None,None,None],
    ["Kühlen [mins]",None,None,15,20,25,None,None,None,None,None,None,None,None,None],
    ["Würzetemp [°C]",None,None,wortTemp-1,wortTemp,wortTemp+1,None,None,None,None,None,None,None,None,None],
    ["Stw [°P]",None,None,og-0.1,og,og+0.1,None,None,None,None,None,None,None,None,None],
    ["pH" ,None,None,mashpH-0.5,mashpH-0.4,mashpH-0.3,None,None,None,None,None,None,None,None,None]

]
#############################ferm sheet
df2 = [
    [None,None,None,None,None,None,None,None,None,None,None,None,None], #ok
    [beerName ,None,None,None,"[BatchNo]",None,None,None,"[Date]",None,None,None,None ,None,None], #ok
    [style,None,og,"°P",abv,"%",fg,"°P",color,"EBC",ibu,"IBU",None,None,None], #ok
    ["FERMENTATION",None,"ist","min","soll","max","Schritt","Beding." ,None,"Temp","Druck","ABFÜLLUNG",None,"Resp:" ,None,None],
    ["Hefemenge",None,None,None,fermYeastAmount,None,1,globals().get('fermStepName1',None),None,globals().get('fermStepTemp1',None),globals().get('fermStepPressure1',None),"Kegs 20l",None,None,None,None],
    ["Gen",None,None,None,None,None,2,globals().get('fermStepName2',None),None,globals().get('fermStepTemp2',None),globals().get('fermStepPressure2',None),"Kegs 50l",None,None,None,None],
    ["Viability",None,None,None,None,None,3,globals().get('fermStepName3',None),None,globals().get('fermStepTemp3',None),globals().get('fermStepPressure3',None),"Flaschen 0.5",None,None,None,None],
    ["Stammwürze",None,None,og-0.1,og,og+0.1,4,globals().get('fermStepName4',None),None,globals().get('fermStepTemp4',None),globals().get('fermStepPressure4',None),"Flaschen 0.3",None,None,None,None],
    ["Restextrakt",None,None,round((fg-0.1),1),fg,round((fg+0.1),1),5,globals().get('fermStepName5',None),None,globals().get('fermStepTemp5',None),globals().get('fermStepPressure5',None),"Direktausschank",None,None,None,None],
    ["pH" ,None,None,round((mashpH-1.5),2),round((mashpH-1.4),2),round((mashpH-1.3),2),6,globals().get('fermStepName6',None),None,globals().get('fermStepTemp6',None),globals().get('fermStepPressure6',None),None,None,None,None,None],
    ["Alkoholgehalt",None,None,round((abv-0.1),1),abv,round((abv+0.1),1),7,globals().get('fermStepName7',None),None,globals().get('fermStepTemp7',None),globals().get('fermStepPressure7',None),None,None,None,None,None],
    [None,"Datum","Zeit","°P","pH","Temp","soll","Druck","set","Truboff","Bemerkungen",None,None,None,None,"Initialen"], #ok
    [1 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [2 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [3 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [4 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [5 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [6 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [7 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [8 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [9 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [10 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [11 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [12 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [13 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [14 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [15 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [16 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [17 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [18 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [19 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [20 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [21 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [22 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [23 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [24 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [25 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [26 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [27 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [28 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [29 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [30 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [31 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [32 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [33 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [34 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [35 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [36 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [37 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [38 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [39 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
    [40 ,None,None,None,None,None,None,None,None,None,None,None,None,None,None]

]

#############################csv export
with open(f'{beerName}_brew.csv', mode='w', newline='') as file:
    writer = csv.writer(file, delimiter=';')
    for row in df:
        # Replace decimal point with comma for floats
        row = [str(value).replace('.', ',') if isinstance(value, float) else value for value in row]
        writer.writerow(row)

with open(f'{beerName}_ferm.csv', mode='w', newline='') as file:
    writer = csv.writer(file, delimiter=';')
    for row in df2:
        # Replace decimal point with comma for floats
        row = [str(value).replace('.', ',') if isinstance(value, float) else value for value in row]
        writer.writerow(row)


#############################xslx fusion
wb = load_workbook(excel_layout)

ws1 = wb['brew']
ws2 = wb['ferm']

start_row = 1
start_column = 1
for row in df:
    for col_num, value in enumerate(row, start=start_column):
        ws1.cell(row=start_row, column=col_num, value=value)
    start_row += 1  # Move to the next row

start_row = 1
start_column = 1
for row in df2:
    for col_num, value in enumerate(row, start=start_column):
        ws2.cell(row=start_row, column=col_num, value=value)
    start_row += 1  # Move to the next row

wb.save(f'{beerName}_brewsheet.xlsx')
